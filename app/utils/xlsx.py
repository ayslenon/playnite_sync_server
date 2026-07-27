from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from sqlmodel import select, Session
from sqlalchemy.orm import selectinload

from app.models import Game


HEADERS = [
    ("Título", 45),
    ("Status", 16),
    ("Plataforma", 20),
    ("Gêneros", 28),
    ("Armazenamento", 20),
    ("Vontade", 10),
    ("Nota", 8),
    ("Replay", 8),
    ("Horas Jogadas", 14),
    ("HLTB Principal", 14),
    ("HLTB Extra", 12),
    ("HLTB 100%", 12),
    ("Finalizado (h)", 14),
    ("Finalizado (data)", 16),
    ("Precisa Testar", 14),
    ("Jogadores", 18),
    ("Tipo Coop", 16),
    ("Tela Coop", 14),
    ("Input Recomendado", 18),
    ("Notas", 40),
    ("Criado em", 22),
    ("Atualizado em", 22),
]

HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Segoe UI", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

STRIPED_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

STATUS_FILLS = {
    "Finalizado": PatternFill(
        start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
    ),
    "Jogando": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "Backlog": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "Abandonado": PatternFill(
        start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
    ),
}


def _format_playtime(seconds: int) -> str:
    if not seconds:
        return ""
    hours = seconds // 3600
    mins = (seconds % 3600) // 60
    if hours > 0:
        return f"{hours}h{mins:02d}min"
    return f"{mins}min"


def _status_val(status: str) -> int:
    order = {"Jogando": 0, "Finalizado": 1, "Backlog": 2, "Abandonado": 3}
    return order.get(status, 99)


def export_games(session: Session) -> BytesIO:
    stmt = (
        select(Game)
        .options(
            selectinload(Game.platform),
            selectinload(Game.storage_device),
            selectinload(Game.genres),
        )
        .order_by(Game.updated_at.desc())
    )
    games = session.exec(stmt).all()

    games.sort(key=lambda g: _status_val(g.gameplay_status))

    wb = Workbook()
    ws = wb.active
    ws.title = "Biblioteca de Jogos"

    for col_idx, (header, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(games) + 1}"
    ws.freeze_panes = "A2"

    for row_idx, game in enumerate(games, 2):
        is_odd = (row_idx % 2) == 0

        status_fill = STATUS_FILLS.get(game.gameplay_status)
        platform_name = game.platform.name if game.platform else ""
        storage_name = game.storage_device.name if game.storage_device else ""
        genre_names = ", ".join(g.name for g in game.genres) if game.genres else ""

        coop_types = game.coop_type_list()

        row_data = [
            game.title,
            game.gameplay_status,
            platform_name,
            genre_names,
            storage_name,
            game.interest_rating,
            game.score or "",
            game.replay_score or "",
            _format_playtime(game.playtime_seconds),
            int(game.hltb_main) if game.hltb_main else "",
            int(game.hltb_main_extra) if game.hltb_main_extra else "",
            int(game.hltb_full) if game.hltb_full else "",
            game.finish_hours or "",
            game.finish_date[:10] if game.finish_date else "",
            "Sim" if game.must_test else "",
            game.coop_players,
            ", ".join(coop_types) if coop_types else "",
            game.coop_screen_type,
            game.input_recommendation,
            game.notes or "",
            game.created_at[:10] if game.created_at else "",
            game.updated_at[:10] if game.updated_at else "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (6, 7, 8, 15) else "left",
                vertical="center",
            )

            if status_fill:
                cell.fill = status_fill
            elif is_odd:
                cell.fill = STRIPED_FILL

        ws.row_dimensions[row_idx].height = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
